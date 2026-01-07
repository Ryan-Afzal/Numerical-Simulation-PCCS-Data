# postProcessing.py
# 
# Ryan Afzal
# 2026-01-07
# 
# Script post-processes the formatted COMSOL data files in Excel workbooks.
# It should load in all files, sort into different types (10 L/D, Centerline, Outlet, etc.), remove extraneous columns, and remove the 0.186 error.
# It should then load the data into a table for each file, then adjust r -> [r], z -> [L/D]
# 
# Finally, it should produce plots, organized into folders, as defined at the end using this data.

import matplotlib.pyplot as plt
import openpyxl
import os.path
import shutil
import sys

profileXMajorTicks2D = [0, 0.5, 1, 1.14]
profileXMinorTicks2D = [0.1, 0.2, 0.3, 0.4, 0.6, 0.7, 0.8, 0.9, 1.1]

profileXMajorTicks3D = [-1.14, -1, -0.5, 0, 0.5, 1, 1.14]
profileXMinorTicks3D = [-1.1, -0.9, -0.8, -0.7, -0.6, -0.4, -0.3, -0.2, -0.1, 0.1, 0.2, 0.3, 0.4, 0.6, 0.7, 0.8, 0.9, 1.1]

axialXMajorTicks = [0, 10, 20, 30, 40, 50, 51.4]
axialXMinorTicks = range(0, 52, 2)

def getProfilePlotTemplate(is3D: bool):
    fig, ax = getGenericPlot()
    
    # Apply settings
    if (is3D):
        ax.set_xlim(xmin=-1.14, xmax=1.14)
        ax.set_xticks(profileXMajorTicks3D, ["-1.14", "-1", "-0.5", "0", "0.5", "1", "1.14"])
        ax.set_xticks(profileXMinorTicks3D, minor=True)
        ax.axvline(-1, ls='--', color='k')
    else:
        ax.set_xlim(xmin=0, xmax=1.14)
        ax.set_xticks(profileXMajorTicks2D, ["0", "0.5", "1", "1.14"])
        ax.set_xticks(profileXMinorTicks2D, minor=True)
    ax.axvline(1, ls='--', color='k')
    

    return fig, ax

def getAxialPlotTemplate():
    fig, ax = getGenericPlot()
    
    # Apply settings
    ax.set_xlim(xmin=0, xmax=51.4)
    ax.set_xticks(axialXMajorTicks, ["0", "10", "20", "30", "40", "50", ""])
    ax.set_xticks(axialXMinorTicks, minor=True)

    return fig, ax

def getGenericPlot():
    fig, ax = plt.subplots(nrows=1, ncols=1)
    
    # Apply any generic settings here
    ax.grid(visible=True, which='major', axis='both')
    ax.grid(visible=True, which='minor', axis='both', alpha=0.5)
    
    return fig, ax

input("Ready, press ENTER to continue:")

print('Beginning Post-Processing...')

# Folder paths
INPUT_FOLDER_PATH = os.getcwd()
OUTPUT_FOLDER_PATH = os.getcwd()

# File extension
FILE_EXT = 'jpeg'

# Ensure input and output folders exist
if (os.path.exists(INPUT_FOLDER_PATH) and os.path.isdir(INPUT_FOLDER_PATH)):
    print(f'\tInput from:\t\'{INPUT_FOLDER_PATH}\'')
else:
    input(f'Input folder \'{INPUT_FOLDER_PATH}\' does not exist.')
    sys.exit()
if (os.path.exists(OUTPUT_FOLDER_PATH) and os.path.isdir(OUTPUT_FOLDER_PATH)):
    print(f'\tOutput to:\t\'{OUTPUT_FOLDER_PATH}\'')
else:
    input(f'Output folder \'{OUTPUT_FOLDER_PATH}\' does not exist.')
    sys.exit()

print()
print()



# Spreadsheet tab labels, which will also be used to generate the data dictionary.
    # File endings, which all end in *.txt. These will be used to generate the data dictionary.
PROFILE_10LD = "Profile_10LD"
PROFILE_20LD = "Profile_20LD"
PROFILE_30LD = "Profile_30LD"
PROFILE_40LD = "Profile_40LD"
PROFILE_50LD = "Profile_50LD"
PROFILE_OUTLET = "Profile_Outlet"

AXIAL_CENTERLINE = "Centerline-Axial"
AXIAL_BL = "Boundary-Layer-Axial"
AXIAL_INNER = "Inner_Edge"
AXIAL_OUTER = "Outer_Edge"

PROFILES_NAMES = [PROFILE_10LD, PROFILE_20LD, PROFILE_30LD, PROFILE_40LD, PROFILE_50LD, PROFILE_OUTLET]
AXIAL_NAMES = [AXIAL_CENTERLINE, AXIAL_BL, AXIAL_INNER, AXIAL_OUTER]

# Dictionaries containing the raw file data, ordered by data source type (PROFILES_NAMES and AXIAL_NAMES).
# 
# formattedDataProfiles => {
#   "Profile_10LD" => {
#        "ME96_case_J" => [...] (Containing the actual raw data, as an array of row entries)
#       }
#   ...
#   }
formattedDataProfiles = {}
formattedDataAxial = {}

# Set up the named dictionary elements.
for name in PROFILES_NAMES:
    formattedDataProfiles[name] = {}
for name in AXIAL_NAMES:
    formattedDataAxial[name] = {}

#contains the header columns, in order, to take
PROFILE_SCHEMA = ["r","T","w","q","k"]
AXIAL_SCHEMA = ["z","T","q","q0"]

SCHEMA_REPLACEMENTS = {
    "r":"y",
    "T":"T (K)",
    "w":"w (m/s)",
    "q":"-(ht.tfluxr) (W/m^2)",
    "q0":"ht.hf1.q0 (W/m^2)",
    "k":"k (m^2/s^2)"
    }

def readSheet(sheet, schema):
    mr = sheet.max_row
    mc = sheet.max_column

    headers = []
    for i in range(1, mc + 1):
        headers.append(str(sheet.cell(row = 1, column = i).value).strip())
    map = {}
    for i in range(0, len(schema)):
        if (headers.count(schema[i]) > 0):
            map[i] = headers.index(schema[i])
        elif(headers.count(SCHEMA_REPLACEMENTS[schema[i]]) > 0):#if there is a replacement, use that
            map[i] = headers.index(SCHEMA_REPLACEMENTS[schema[i]])
    
    out_array = []
    for r in range(2, mr + 1):
        row = []
        for i in range(0, len(schema)):
            if (map.__contains__(i)):
                val = str(sheet.cell(row = r, column = (map[i] + 1)).value)
                row.append('NaN' if val == '#NUM!' else val)
            else:
                row.append("0")
        out_array.append(row)
    return out_array

# Search through each Excel workbook and read in data from each sheet into the dictionaries.
for file in os.scandir(INPUT_FOLDER_PATH):
    if (file.is_dir()):
        continue
    if (file.name.endswith(".xlsx")):
        print(f'\tReading workbook \'{file.name}\'...')
        caseName = file.name[:file.name.index(".xlsx")]
        workbook = openpyxl.load_workbook(file.name)
        for sheet in workbook.worksheets:
            sheetName = sheet.title
            found = False
            print(f'\t\tReading sheet \'{sheetName}\'...')
            for name in PROFILES_NAMES:# Check against profiles
                if (sheetName == name):
                    found = True
                    readArray = readSheet(sheet, PROFILE_SCHEMA)
                    formattedDataProfiles[name][caseName] = readArray
            for name in AXIAL_NAMES:# Check against axial
                if (sheetName == name):
                    found = True
                    readArray = readSheet(sheet, AXIAL_SCHEMA)
                    formattedDataAxial[name][caseName] = readArray
            if (found):
                print(f'\t\t\tDone.')
            else:
                print(f'\t\t\tUnknown sheet, not read!')
        print(f'\t\tDone.')
    else:
        print(f'\tSkipping non-data file \'{file.name}\'.')
print('Done reading files.')

print()
print('Loaded data (Profile):')
for key in formattedDataProfiles.keys():
    print(f'\t{key}:')
    for key2 in formattedDataProfiles[key].keys():
        print(f'\t\t{key2}:')
print()
print('Loaded data (Axial):')
for key in formattedDataAxial.keys():
    print(f'\t{key}:')
    for key2 in formattedDataAxial[key].keys():
        print(f'\t\t{key2}:')
print()
print('Sorting data...')

# Sort profile data
for name in formattedDataProfiles:
    for key in formattedDataProfiles[name]:
        entry = formattedDataProfiles[name][key]
        formattedDataProfiles[name][key].sort(key=lambda row: float(row[0]))

# Remove 0.0186 error and duplicates
# (When COMSOL exports data at the fluid-solid boundary (r=0.0186m), it doesn't ensure they are in the correct order)
for name in formattedDataProfiles:
    for key in formattedDataProfiles[name]:
        entry = formattedDataProfiles[name][key]
        rowIndex = 0
        while (rowIndex < len(entry) - 1):
            if (entry[rowIndex][0] == entry[rowIndex + 1][0]):
                if (entry[rowIndex][1] == entry[rowIndex + 1][1]):
                    del entry[rowIndex]
                    continue
                elif (entry[rowIndex][0][0] == '-'):
                    if (float(entry[rowIndex][1]) < float(entry[rowIndex + 1][1])):
                        swap = entry[rowIndex + 1][1]
                        entry[rowIndex + 1][1] = entry[rowIndex][1]
                        entry[rowIndex][1] = swap
                else:
                    if (float(entry[rowIndex][1]) > float(entry[rowIndex + 1][1])):
                        swap = entry[rowIndex + 1][1]
                        entry[rowIndex + 1][1] = entry[rowIndex][1]
                        entry[rowIndex][1] = swap
            rowIndex += 1

# Sort axial data
for name in formattedDataAxial:
    for key in formattedDataAxial[name]:
        formattedDataAxial[name][key].sort(key=lambda row: float(row[0]))

# Filter function for HF data
def applyHeatFluxFilter(dataset, qIndex):
    for caseName in dataset:
        data = dataset[caseName]
        for i in range(1, len(data) - 1):
            data[i][qIndex] = str(0.25*float(data[i - 1][qIndex]) + 0.5*float(data[i][qIndex]) + 0.25*float(data[i - 1][qIndex]))

# Finder for max profile deviation
def findMaxDeviation(plotVar: str, source: str, case1: str, case2: str, probePoints):
    probeValues = []

    if (PROFILES_NAMES.count(source) > 0):
        case1Data = formattedDataProfiles[source][case1]
        case2Data = formattedDataProfiles[source][case2]
        plotVarIndex = PROFILE_SCHEMA.index(plotVar)
    else:
        case1Data = formattedDataAxial[source][case1]
        case2Data = formattedDataAxial[source][case2]
        plotVarIndex = AXIAL_SCHEMA.index(plotVar)

    if (probePoints[0] < 0):
        arr = case1Data.copy()
        arr2 = []
        for item in arr:
            if (item[0] != "0"):
                arr2.append(item)
                if (arr2[-1][0].startswith('-')):
                    arr2[-1][0] = arr2[-1][0][1:]
                else:
                    arr2[-1][0] = f'-{arr2[-1][0]}'
        for item in arr2:
            arr.insert(0,item)
        case1Data = arr

    for point in probePoints:
        case1Index = 0
        for i in range(len(case1Data) - 1):
            if (float(case1Data[i + 1][0]) > point):
                break
            case1Index = i
        case2Index = 0
        for i in range(len(case2Data) - 1):
            if (float(case2Data[i + 1][0]) > point):
                break
            case2Index = i
        
        case1_x0 = float(case1Data[case1Index][0])
        case1_x1 = float(case1Data[case1Index + 1][0])
        case1_y0 = float(case1Data[case1Index][plotVarIndex])
        case1_y1 = float(case1Data[case1Index + 1][plotVarIndex])
        if (case1_x0 == case1_x1):
            probeValues.append([-1, -1, -1])
            continue
        case1Y = (case1_y0*(case1_x1 - point) + case1_y1*(point - case1_x0)) / (case1_x1 - case1_x0)

        case2_x0 = float(case2Data[case2Index][0])
        case2_x1 = float(case2Data[case2Index + 1][0])
        case2_y0 = float(case2Data[case2Index][plotVarIndex])
        case2_y1 = float(case2Data[case2Index + 1][plotVarIndex])
        if (case2_x0 == case2_x1):
            probeValues.append([-1, -1, -1])
            continue
        case2Y = (case2_y0*(case2_x1 - point) + case2_y1*(point - case2_x0)) / (case2_x1 - case2_x0)
        
        probeValues.append([case1Y, case2Y, abs(case2Y - case1Y), [case1_x0, case1_x1, case1_y0, case1_y1], [case2_x0, case2_x1, case2_y0, case2_y1]])
    
    maxIndex = 0
    for i in range(len(probeValues)):
        if (probeValues[i][2] > probeValues[maxIndex][2]):
            maxIndex = i
    
    maxX = probePoints[maxIndex]
    maxY = probeValues[maxIndex][2]
    maxYP = probeValues[maxIndex][2] / max(probeValues[maxIndex][0], probeValues[maxIndex][1])

    return maxX, maxY, maxYP

print('Done!')
print()

input(f'Press ENTER to continue:')

print()
print()
print(f'Post-Processing Data...')
print()
print()

# Post-processing definitions

R = 0.0186

XFACTOR_R = R
XFACTOR_LD = 2*R

R_AXIS_TITLE = "Normalized Radial Position (r/R)"
Z_AXIS_TITLE = "Vertical Position (L/D)"

TEMPERATURE = "T"
W_VELOCITY = "w"
HEAT_FLUX = "q"
TKE = "k"

Y_FACTOR = {
    TEMPERATURE: 1,
    W_VELOCITY: 1,
    HEAT_FLUX: 1E-3,
    TKE: 1
}

Y_TITLE = {
    TEMPERATURE: "Temperature (K)",
    W_VELOCITY: "Vertical Velocity (m/s)",
    HEAT_FLUX: "Inward Heat Flux (kW/m2)",
    TKE: "Turbulent Kinetic Energy (m^2/s^2)"
}

Y_LIM = {
    TEMPERATURE: [
        [300, 310, 320, 330, 340, 350, 360, 370],
        range(300, 372, 2)
    ],
    W_VELOCITY: [
        [0, 0.1, 0.2, 0.3, 0.4],
        [0, 0.02, 0.04, 0.06, 0.08, 0.12, 0.14, 0.16, 0.18, 0.22, 0.24, 0.26, 0.28, 0.32, 0.34, 0.36, 0.38]
    ],
    HEAT_FLUX: [
        range(0, 80, 10),
        range(0, 82, 2)
    ],
    TKE: [
        [1E-4, 1E-3],
        [1E-4, 2E-4, 5E-4, 1E-3]
    ]
}

NRC = "Benchmark"
C1 = "Case_c1"
C2 = "Case_c2"
C3 = "Case_c3"
C4 = "Case_c4"
CA = "Case_A"
CB = "Case_B"
CC = "Case_C"
CD = "Case_D"
CE = "Case_E"
CF = "Case_F"
CG = "Case_G"
CH = "Case_H"
CI = "Case_I"
CJ = "Case_J"
CK = "Case_K"
CL = "Case_L"
CM = "Case_M"
CN = "Case_N"

PROBE_POINTS_2D_W = []
i = 0
while (i < 0.9):
    PROBE_POINTS_2D_W.append(i*R)
    i += 0.01
    
PROBE_POINTS_2D_T = []
i = 0
while (i < 1.14):
    PROBE_POINTS_2D_T.append(i*R)
    i += 0.01

PROBE_POINTS_3D_W = []
i = 0
while (i < 0.9):
    if (i != 1 and i != -1):
        PROBE_POINTS_3D_W.append(i*R)
        PROBE_POINTS_3D_W.insert(0, -i*R)
    i += 0.01
    
PROBE_POINTS_3D_T = []
i = 0
while (i < 1.14):
    if (i != 1 and i != -1):
        PROBE_POINTS_3D_T.append(i*R)
        PROBE_POINTS_3D_T.insert(0, -i*R)
    i += 0.01

outputDir = ""

def plot(title: str, plotVar: str, dataSource, cases: list, caseLabels: list, is3D = False, isLog = False, yTicks = None):
    singleSource = type(dataSource) is not list
    if (singleSource):# convert dataSource to list
        dataSource = [dataSource]
        if (len(cases) != len(caseLabels)):# ensure cases match to labels
            input(f'Cannot plot: # of cases ({len(cases)}) != # of labels ({len(caseLabels)})')
            return
    else:
        if (len(cases) != 1):# if multi-source, len(cases) must be 1
            input(f'\tCannot plot: too many arguments!')
            return
        if (len(dataSource) != len(caseLabels)):# ensure sources match to labels
            input(f'Cannot plot: # of sources ({len(dataSource)}) != # of labels ({len(caseLabels)})')
            return
    
    isProfile = PROFILES_NAMES.count(dataSource[0]) > 0
    firstColIndex = 0
    if (isProfile):# profile plot
        data = formattedDataProfiles
        secondColIndex = PROFILE_SCHEMA.index(plotVar)

        fig, ax = getProfilePlotTemplate(is3D)

        ax.set_xlabel(R_AXIS_TITLE)
        xFactor = XFACTOR_R

        print(f'Creating profile plot {title}...')
    else:# axial plot
        data = formattedDataAxial
        secondColIndex = AXIAL_SCHEMA.index(plotVar)

        fig, ax = getAxialPlotTemplate()
        ax.set_xlabel(Z_AXIS_TITLE)
        xFactor = XFACTOR_LD

        print(f'Creating axial plot {title}...')

    if (yTicks is None):# set plot axes
        yTicks = Y_LIM[plotVar]
    ax.set_yticks(yTicks[0])
    ax.set_yticks(yTicks[1], minor=True)
    ax.set_ylim(yTicks[0][0], yTicks[0][-1])
    if (isLog):
        ax.set_yscale('log')

    # set plot title and y-axis label
    ax.set_title(title)
    ax.set_ylabel(Y_TITLE[plotVar])

    # load data into plot
    for dIndex in range(len(dataSource)):
        for cIndex in range(len(cases)):
            xAxis = []
            yAxis = []
            entry = data[dataSource[dIndex]][cases[cIndex]]
            clone = isProfile and is3D and entry[0][firstColIndex][0] != '-'    
            for r in range(len(entry)):
                x = float(entry[r][firstColIndex]) / xFactor
                y = float(entry[r][secondColIndex]) * Y_FACTOR[plotVar]
                xAxis.append(x)
                yAxis.append(y)
                if (clone):
                    xAxis.insert(0, -x)
                    yAxis.insert(0, y)
            
            if (singleSource):
                labelIndex = cIndex
            else:
                labelIndex = dIndex
            if (type(caseLabels[labelIndex]) is list):
                label = caseLabels[labelIndex][0]
                format = caseLabels[labelIndex][1]
                ax.plot(xAxis, yAxis, format, label=label)
            else:
                ax.plot(xAxis, yAxis, label=caseLabels[labelIndex])

    ax.legend()
    fileTitle = title.replace('/', '-').replace('\\', '-')
    if (not os.path.exists(os.path.join(OUTPUT_FOLDER_PATH, outputDir))):
        os.makedirs(os.path.join(OUTPUT_FOLDER_PATH, outputDir))
    fileTitle = os.path.join(OUTPUT_FOLDER_PATH, outputDir, fileTitle)
    if (os.path.exists(f'{fileTitle}.{FILE_EXT}')):
        fileTitle += '_v2'
        count = 2
        while (os.path.exists(f'{fileTitle}.{FILE_EXT}')):
            count += 1
            fileTitle = fileTitle[:-1] + str(count)
    fig.savefig(f'{fileTitle}.{FILE_EXT}', bbox_inches='tight', dpi=300)

    print(f'\tDone! Saved file \'{fileTitle}.{FILE_EXT}\'')

#
# Plot Definitions
#


def calculateDeviation():
    print(f'Deviation for MX Cases...')
    mX, mY, mP = findMaxDeviation(W_VELOCITY, PROFILE_50LD, CB, CC, PROBE_POINTS_2D_W)
    print(f'\tW 50 L/D, B/C; X: {mX}; Y: {mY}, %: {mP * 100}')
    mX, mY, mP = findMaxDeviation(W_VELOCITY, PROFILE_50LD, CJ, CK, PROBE_POINTS_2D_W)
    print(f'\tW 50 L/D, J/K; X: {mX}; Y: {mY}, %: {mP * 100}')
    mX, mY, mP = findMaxDeviation(W_VELOCITY, PROFILE_50LD, CL, CM, PROBE_POINTS_2D_W)
    print(f'\tW 50 L/D, L/M; X: {mX}; Y: {mY}, %: {mP * 100}')
    mX, mY, mP = findMaxDeviation(TEMPERATURE, PROFILE_50LD, CB, CC, PROBE_POINTS_2D_T)
    print(f'\tT 50 L/D, B/C; X: {mX}; Y: {mY}, %: {mP * 100}')
    mX, mY, mP = findMaxDeviation(TEMPERATURE, PROFILE_50LD, CJ, CK, PROBE_POINTS_2D_T)
    print(f'\tT 50 L/D, J/K; X: {mX}; Y: {mY}, %: {mP * 100}')
    mX, mY, mP = findMaxDeviation(TEMPERATURE, PROFILE_50LD, CL, CM, PROBE_POINTS_2D_T)
    print(f'\tT 50 L/D, L/M; X: {mX}; Y: {mY}, %: {mP * 100}')

    print(f'Deviation for 2D vs 3D...')
    mX, mY, mP = findMaxDeviation(W_VELOCITY, PROFILE_30LD, CB, CE, PROBE_POINTS_2D_W)
    print(f'\tW 30 L/D, B/E; X: {mX}; Y: {mY}, %: {mP * 100}')
    mX, mY, mP = findMaxDeviation(W_VELOCITY, PROFILE_50LD, CB, CE, PROBE_POINTS_2D_W)
    print(f'\tW 50 L/D, B/E; X: {mX}; Y: {mY}, %: {mP * 100}')

    print(f'Deviation for 2D vs 3D MX...')
    mX, mY, mP = findMaxDeviation(W_VELOCITY, PROFILE_50LD, CB, CE, PROBE_POINTS_2D_W)
    print(f'\tW 50 L/D, B/E; X: {mX}; Y: {mY}, %: {mP * 100}')
    mX, mY, mP = findMaxDeviation(W_VELOCITY, PROFILE_50LD, CC, CF, PROBE_POINTS_2D_W)
    print(f'\tW 50 L/D, C/F; X: {mX}; Y: {mY}, %: {mP * 100}')
    #mX, mY, mP = findMaxDeviation(W_VELOCITY, PROFILE_50LD, CE, CF, PROBE_POINTS_2D_W)
    #print(f'\tW 50 L/D, E/F; X: {mX}; Y: {mY}, %: {mP * 100}')

def plotMX3D():
    global outputDir
    outputDir = "MX 3D"
    plot("Temperature Profiles (E)", TEMPERATURE, [PROFILE_10LD, PROFILE_30LD, PROFILE_50LD], [CE], [
        "10 L/D",
        "30 L/D",
        "50 L/D"
    ], is3D=True)
    plot("Velocity Profiles (E)", W_VELOCITY, [PROFILE_10LD, PROFILE_30LD, PROFILE_50LD], [CE], [
        "10 L/D",
        "30 L/D",
        "50 L/D"
    ], is3D=True)
    plot("Temperature Profiles (F)", TEMPERATURE, [PROFILE_10LD, PROFILE_30LD, PROFILE_50LD], [CF], [
        "10 L/D",
        "30 L/D",
        "50 L/D"
    ], is3D=True)
    plot("Velocity Profiles (F)", W_VELOCITY, [PROFILE_10LD, PROFILE_30LD, PROFILE_50LD], [CF], [
        "10 L/D",
        "30 L/D",
        "50 L/D"
    ], is3D=True)
    outputDir = ""

def plotME94Series():
    global outputDir
    outputDir = "Benchmark Comparison"
    plot("Outlet Temperature Profiles", TEMPERATURE, PROFILE_OUTLET, [C1, C2, C3, C4], [
        "2D Normal (c1)", 
        "3D Normal (c2)", 
        "2D Refined (c3)",
        "2D k-ε Realizable (c4)"
    ], is3D=True)
    plot("Outlet Temperature Profiles", TEMPERATURE, PROFILE_OUTLET, [C3, NRC], [
        "2D Refined (c3)",
        "NRC Benchmark"
    ], is3D=True)
    plot("Turbulent Kinetic Energy Profiles (50 L/D)", TKE, PROFILE_50LD, [C1, C2, C3, C4], [
        "2D Normal (c1)", 
        "3D Normal (c2)", 
        "2D Refined (c3)",
        "2D k-ε Realizable (c4)"
    ], is3D=True, isLog=True)
    outputDir = ""

def applyFilters():
    global outputDir
    outputDir = "Filter Comparison"
    plot("Inward Heat Flux Before Filter (Inner Wall)", HEAT_FLUX, AXIAL_INNER, [CB, CC, CN, CG, CH, CI], [
        ["Constant HF (B)", '-'],
        ["Constant HF (C)", '--'],
        ["Constant HF (N)", ':'],
        ["Variable HTC (G)", '-'],
        ["Variable HTC (H)", '--'],
        ["Constant HTC (I)", ':']
    ])
    plot("Inward Heat Flux Before Filter (Boundary Layer)", HEAT_FLUX, AXIAL_BL, [CB, CC, CN, CG, CH, CI], [
        ["Constant HF (B)", '-'],
        ["Constant HF (C)", '--'],
        ["Constant HF (N)", ':'],
        ["Variable HTC (G)", '-'],
        ["Variable HTC (H)", '--'],
        ["Constant HTC (I)", ':']
    ])

    applyHeatFluxFilter(formattedDataAxial[AXIAL_INNER], 2)
    applyHeatFluxFilter(formattedDataAxial[AXIAL_BL], 2)

    plot("Inward Heat Flux After Filter (Inner Wall)", HEAT_FLUX, AXIAL_INNER, [CB, CC, CN, CG, CH, CI], [
        ["Constant HF (B)", '-'],
        ["Constant HF (C)", '--'],
        ["Constant HF (N)", ':'],
        ["Variable HTC (G)", '-'],
        ["Variable HTC (H)", '--'],
        ["Constant HTC (I)", ':']
    ])
    plot("Inward Heat Flux After Filter (Boundary Layer)", HEAT_FLUX, AXIAL_BL, [CB, CC, CN, CG, CH, CI], [
        ["Constant HF (B)", '-'],
        ["Constant HF (C)", '--'],
        ["Constant HF (N)", ':'],
        ["Variable HTC (G)", '-'],
        ["Variable HTC (H)", '--'],
        ["Constant HTC (I)", ':']
    ])
    outputDir = ""

def plotMeshComparison():
    global outputDir
    outputDir = "Mesh Comparison"
    plot("Temperature Profiles (50 L/D)", TEMPERATURE, PROFILE_50LD, [CA, CB, CN], [
        "2D Normal (A)", 
        "2D Refined (B)", 
        "2D Extra-Refined (N)"
    ])
    plot("Velocity Profiles (50 L/D)", W_VELOCITY, PROFILE_50LD, [CA, CB, CN], [
        "2D Normal (A)", 
        "2D Refined (B)", 
        "2D Extra-Refined (N)"
    ])
    plot("Inward Heat Flux Profiles (50 L/D)", HEAT_FLUX, PROFILE_50LD, [CA, CB, CN], [
        "2D Normal (A)", 
        "2D Refined (B)", 
        "2D Extra-Refined (N)"
    ])
    plot("Inward Heat Flux (Boundary Layer)", HEAT_FLUX, AXIAL_BL, [CA, CB, CN], [
        "2D Normal (A)", 
        "2D Refined (B)", 
        "2D Extra-Refined (N)"
    ])
    plot("Temperature Profiles (50 L/D)", TEMPERATURE, PROFILE_50LD, [CD, CE], [
        "3D Normal (D)",
        "3D Refined (E)"
    ], is3D=True)
    outputDir = ""

def plot2D3DComparison():
    global outputDir
    outputDir = "2D vs 3D Comparison"
    plot("Temperature Profiles (E)", TEMPERATURE, [PROFILE_10LD, PROFILE_30LD, PROFILE_50LD], [CE], [
        "10 L/D",
        "30 L/D",
        "50 L/D"
    ], is3D=True)
    plot("Temperature Profiles (F)", TEMPERATURE, [PROFILE_10LD, PROFILE_30LD, PROFILE_50LD], [CF], [
        "10 L/D",
        "30 L/D",
        "50 L/D"
    ], is3D=True)
    plot("2D vs 3D Temperature Profiles (30 L/D)", TEMPERATURE, PROFILE_30LD, [CB, CE], [
        "2D (B)",
        "3D (E)"
    ], is3D=True)
    plot("2D vs 3D Velocity Profiles (30 L/D)", W_VELOCITY, PROFILE_30LD, [CB, CE], [
        "2D (B)",
        "3D (E)"
    ], is3D=True)
    plot("2D vs 3D Temperature Profiles (50 L/D)", TEMPERATURE, PROFILE_50LD, [CB, CE], [
        "2D (B)",
        "3D (E)"
    ], is3D=True)
    plot("2D vs 3D Velocity Profiles (50 L/D)", W_VELOCITY, PROFILE_50LD, [CB, CE], [
        "2D (B)",
        "3D (E)"
    ], is3D=True)
    outputDir = ""

def plotThermalComparison():
    global outputDir
    outputDir = "Thermal BC Comparison"
    plot("Temperature Profiles (30 L/D)", TEMPERATURE, PROFILE_30LD, [CB, CC, CG, CH, CI], [
    ["Constant HF (B)", '-'],
    ["Constant HF (C)", '--'],
    ["Variable HTC (G)", '-'],
    ["Variable HTC (H)", '--'],
    ["Constant HTC (I)", ':']
    ])
    plot("Temperature Profiles (50 L/D)", TEMPERATURE, PROFILE_50LD, [CB, CC, CG, CH, CI], [
    ["Constant HF (B)", '-'],
    ["Constant HF (C)", '--'],
    ["Variable HTC (G)", '-'],
    ["Variable HTC (H)", '--'],
    ["Constant HTC (I)", ':']
    ])
    plot("Inward Heat Flux Profiles (50 L/D)", HEAT_FLUX, PROFILE_50LD, [CB, CC, CG, CH, CI], [
    ["Constant HF (B)", '-'],
    ["Constant HF (C)", '--'],
    ["Variable HTC (G)", '-'],
    ["Variable HTC (H)", '--'],
    ["Constant HTC (I)", ':']
    ])

    plot("Temperature (Outer Wall)", TEMPERATURE, AXIAL_OUTER, [CB, CC, CG, CH, CI], [
    ["Constant HF (B)", '-'],
    ["Constant HF (C)", '--'],
    ["Variable HTC (G)", '-'],
    ["Variable HTC (H)", '--'],
    ["Constant HTC (I)", ':']
    ])
    plot("Inward Heat Flux (Outer Wall)", HEAT_FLUX, AXIAL_OUTER, [CB, CC, CG, CH, CI], [
    ["Constant HF (B)", '-'],
    ["Constant HF (C)", '--'],
    ["Variable HTC (G)", '-'],
    ["Variable HTC (H)", '--'],
    ["Constant HTC (I)", ':']
    ])
    plot("Temperature (Boundary Layer)", TEMPERATURE, AXIAL_BL, [CB, CC, CG, CH, CI], [
    ["Constant HF (B)", '-'],
    ["Constant HF (C)", '--'],
    ["Variable HTC (G)", '-'],
    ["Variable HTC (H)", '--'],
    ["Constant HTC (I)", ':']
    ])
    plot("Inward Heat Flux (Boundary Layer)", HEAT_FLUX, AXIAL_BL, [CB, CC, CG, CH, CI], [
    ["Constant HF (B)", '-'],
    ["Constant HF (C)", '--'],
    ["Variable HTC (G)", '-'],
    ["Variable HTC (H)", '--'],
    ["Constant HTC (I)", ':']
    ])
    outputDir = ""

def plotMX2D3D():
    global outputDir
    outputDir = "MX 2D vs 3D"
    plot("2D vs 3D Temperature Profiles (50 L/D)", TEMPERATURE, PROFILE_50LD, [CB, CC, CE, CF], [
        ["2D Incompressible (B)", '-'],
        ["2D Compressible (C)", '--'],
        ["3D Incompressible (E)", '-'],
        ["3D Compressible (F)", '--']
    ], is3D=True)
    plot("2D vs 3D Velocity Profiles (50 L/D)", W_VELOCITY, PROFILE_50LD, [CB, CC, CE, CF], [
        ["2D Incompressible (B)", '-'],
        ["2D Compressible (C)", '--'],
        ["3D Incompressible (E)", '-'],
        ["3D Compressible (F)", '--']
    ], is3D=True)
    outputDir = ""

def plotMXComparison():
    global outputDir
    outputDir = "MX Comparison"
    plot("Temperature Profiles (50 L/D)", TEMPERATURE, PROFILE_50LD, [CB, CC], [
    "0.33 kg/s Incompressible (B)",
    "0.33 kg/s Compressible (C)"
    ])
    plot("Velocity Profiles (50 L/D)", W_VELOCITY, PROFILE_50LD, [CB, CC], [
    "0.33 kg/s Incompressible (B)",
    "0.33 kg/s Compressible (C)"
    ])
    plot("Inward Heat Flux Profiles (50 L/D)", HEAT_FLUX, PROFILE_50LD, [CB, CC], [
    "0.33 kg/s Incompressible (B)",
    "0.33 kg/s Compressible (C)"
    ])
    plot("Temperature Profiles (50 L/D)", TEMPERATURE, PROFILE_50LD, [CJ, CK], [
    "0.28 kg/s Incompressible (J)",
    "0.28 kg/s Compressible (K)"
    ])
    plot("Velocity Profiles (50 L/D)", W_VELOCITY, PROFILE_50LD, [CJ, CK], [
    "0.28 kg/s Incompressible (J)",
    "0.28 kg/s Compressible (K)"
    ])
    plot("Inward Heat Flux Profiles (50 L/D)", HEAT_FLUX, PROFILE_50LD, [CJ, CK], [
    "0.28 kg/s Incompressible (J)",
    "0.28 kg/s Compressible (K)"
    ])
    plot("Temperature Profiles (50 L/D)", TEMPERATURE, PROFILE_50LD, [CL, CM], [
    "0.37 kg/s Incompressible (L)",
    "0.37 kg/s Compressible (M)"
    ])
    plot("Velocity Profiles (50 L/D)", W_VELOCITY, PROFILE_50LD, [CL, CM], [
    "0.37 kg/s Incompressible (L)",
    "0.37 kg/s Compressible (M)"
    ], yTicks=[
        [0, 0.1, 0.2, 0.3, 0.4, 0.5],
        [0, 0.02, 0.04, 0.06, 0.08, 0.12, 0.14, 0.16, 0.18, 0.22, 0.24, 0.26, 0.28, 0.32, 0.34, 0.36, 0.38, 0.42, 0.44, 0.46, 0.48]
    ])
    outputDir = ""

# 
# Run functions
# 

# Apply HF data filter
#
applyFilters()

# Deviation calcs
#
calculateDeviation()

# Plots
#
plotMX3D()
plotME94Series()
plotMeshComparison()
plot2D3DComparison()
plotThermalComparison()
plotMX2D3D()
plotMXComparison()

#
# Finished
#

# Move files to output folder
print(f'Moving files to output folder...')
for file in os.scandir(os.getcwd()):
    if (not file.is_dir() and file.name.endswith(f'.{FILE_EXT}')):
        shutil.move(file, os.path.join(OUTPUT_FOLDER_PATH, file.name))
print(f'\tDone!')

print()
print()
input(f'Finished! Press ENTER to exit:')